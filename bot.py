#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import io
import json
import asyncio
from datetime import datetime, timedelta
from typing import List, Tuple

import aiosqlite
import openpyxl
from openpyxl.styles import Font

from PIL import Image
from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery,
    BufferedInputFile
)
from aiogram.filters import CommandStart, StateFilter 
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from dotenv import load_dotenv

# Google AI
import google.generativeai as genai

# === Настройки ===
load_dotenv()

# ⚠️ ВСТАВЬТЕ СЮДА ВАШИ КЛЮЧИ (или используйте .env файл)
BOT_TOKEN = os.getenv("BOT_TOKEN", "7625061072:AAH_5PlnjKRY2sepDrxfU066PcPDMA5vf9Q")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "AIzaSyBfj-lEGVAnSsCcqEbtjMNmPqLeC_ReWcM")

if not GEMINI_API_KEY:
    print("❌ ОШИБКА: Не найден GEMINI_API_KEY.")
    exit()

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-2.0-flash')

DB_PATH = os.getenv("DB_PATH", "expenses.db")

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Временное хранилище для редактируемых чеков (user_id -> list)
PENDING_GOODS: dict[int, List[Tuple[str, float]]] = {}

# Список кнопок меню
MENU_BUTTONS = ["➕ Добавить трату", "📷 Чек (фото)", "📋 Список трат", 
                "📊 Статистика", "🗑 Удалить историю", "📆 Отчёт за период"]

# === Состояния (FSM) ===
class ExpenseState(StatesGroup):
    waiting_for_manual_input = State() # Ожидание ввода "Товар цена"
    waiting_for_days = State()         # Ожидание числа дней для отчета

class OCRState(StatesGroup):
    viewing_list = State()         # Просмотр списка (чтобы хранить id сообщения)
    waiting_for_line_number = State()  # Ожидание номера строки
    waiting_for_new_data = State()     # Ожидание новых данных строки

# === Клавиатуры ===

def get_main_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Добавить трату"), KeyboardButton(text="📷 Чек (фото)")],
            [KeyboardButton(text="📋 Список трат"), KeyboardButton(text="📊 Статистика")],
            [KeyboardButton(text="🗑 Удалить историю"), KeyboardButton(text="📆 Отчёт за период")]
        ],
        resize_keyboard=True
    )

def get_cancel_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Отмена", callback_data="cancel_action")]
    ])

def get_ocr_kb():
    """Меню управления распознанным чеком"""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Ред. строку", callback_data="ocr_edit"),
         InlineKeyboardButton(text="❌ Удалить строку", callback_data="ocr_delete")],
        [InlineKeyboardButton(text="✅ Сохранить всё", callback_data="ocr_save"),
         InlineKeyboardButton(text="🚫 Сброс", callback_data="ocr_cancel")]
    ])

def get_report_format_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Сообщением", callback_data="rep_text")],
        [InlineKeyboardButton(text="📊 Excel файл", callback_data="rep_excel")],
        [InlineKeyboardButton(text="🔙 Отмена", callback_data="cancel_action")]
    ])

def get_confirm_delete_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔥 Да, удалить", callback_data="confirm_delete_history")],
        [InlineKeyboardButton(text="🔙 Нет", callback_data="cancel_action")]
    ])

# === БД ===
async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                price REAL,
                date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.commit()

# === AI Логика ===
async def recognize_receipt_with_ai(image_bytes: bytes) -> List[Tuple[str, float]]:
    try:
        img = Image.open(io.BytesIO(image_bytes))
        prompt = """
        Ты - система OCR для чеков. Извлеки список товаров и цены.
        Верни ТОЛЬКО JSON формат: [["Название", 100.0], ["Название 2", 50.0]]
        Игнорируй скидки и промежуточные итоги. Исправляй названия на человекочитаемые.
        """
        response = await model.generate_content_async([prompt, img])
        text_resp = response.text.strip()
        if text_resp.startswith("```"):
            text_resp = text_resp.replace("```json", "").replace("```", "").strip()
        data = json.loads(text_resp)
        result = []
        for item in data:
            if len(item) == 2:
                try:
                    price_clean = str(item[1]).replace(',', '.').replace(' ', '')
                    result.append((str(item[0]).strip(), float(price_clean)))
                except ValueError:
                    continue
        return result
    except Exception as e:
        print(f"AI Error: {e}")
        return []

# === Хелпер: Проверка на нажатие меню ===
async def check_menu_break(message: Message, state: FSMContext) -> bool:
    if message.text in MENU_BUTTONS:
        await state.clear()
        await message.answer("⚠️ Действие прервано. Выберите пункт меню.", reply_markup=get_main_kb())
        return True
    return False

# === Хелпер: Удаление сообщений ===
async def try_delete(bot: Bot, chat_id: int, message_id: int):
    try:
        await bot.delete_message(chat_id, message_id)
    except:
        pass

# === Хендлеры ===

@dp.message(CommandStart())
async def start_cmd(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Привет! Я веду учет расходов.", reply_markup=get_main_kb())

@dp.callback_query(F.data == "cancel_action")
async def global_cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    await callback.message.answer("❌ Отменено.", reply_markup=get_main_kb())

# --- 1. Добавление трат (Ручное) ---

@dp.message(F.text == "➕ Добавить трату")
async def manual_add_prompt(message: Message, state: FSMContext):
    await state.set_state(ExpenseState.waiting_for_manual_input)
    await message.answer("Введи: <b>Товар Цена</b> (например: Хлеб 50)", parse_mode="HTML", reply_markup=get_cancel_kb())

@dp.message(ExpenseState.waiting_for_manual_input)
async def manual_add(message: Message, state: FSMContext):
    if await check_menu_break(message, state): return

    try:
        parts = message.text.rsplit(maxsplit=1)
        if len(parts) < 2: raise ValueError
        name, price = parts[0].strip(), float(parts[1].replace(',', '.'))
        
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT INTO expenses (name, price) VALUES (?, ?)", (name, price))
            await db.commit()
        
        # Исправление: Сразу очищаем состояние, чтобы бот не ждал следующую строку
        await state.clear()
        await message.answer(f"✅ Добавлено: {name} — {price} ₽\n(Бот готов к новым командам)")
        
    except ValueError:
        await message.answer("⚠️ Ошибка формата. Пример: <i>Молоко 90</i>", parse_mode="HTML")

# Быстрый ввод (без кнопки)
# Быстрый ввод (без кнопки)
# Добавили StateFilter(None), чтобы он не перехватывал ввод при редактировании чека
@dp.message(F.text.regexp(r"^(?=.*[^\d\s])(.+)\s(\d+[.,]?\d*)$"), StateFilter(None)) 
async def fast_manual_add(message: Message):
    try:
        parts = message.text.rsplit(maxsplit=1)
        name, price = parts[0].strip(), float(parts[1].replace(',', '.'))
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT INTO expenses (name, price) VALUES (?, ?)", (name, price))
            await db.commit()
        await message.answer(f"✅ {name}: {price}")
    except: pass
# --- 2. Отчеты ---

@dp.message(F.text == "📆 Отчёт за период")
async def ask_report_days(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("За сколько дней нужен отчет? Введи число:", reply_markup=get_cancel_kb())
    await state.set_state(ExpenseState.waiting_for_days)

@dp.message(ExpenseState.waiting_for_days)
async def report_days_received(message: Message, state: FSMContext):
    if await check_menu_break(message, state): return

    if not message.text.isdigit():
        await message.answer("Пожалуйста, введи целое число.")
        return

    days = int(message.text)
    await state.update_data(days=days)
    await message.answer(f"Формат отчета за {days} дней?", reply_markup=get_report_format_kb())

@dp.callback_query(F.data == "rep_text")
async def report_text(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    days = data.get("days", 7)
    await state.clear()

    start_date = datetime.now() - timedelta(days=days)
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT name, price, date FROM expenses WHERE date >= ? ORDER BY date DESC", (start_date,)) as cur:
            rows = await cur.fetchall()

    if not rows:
        await callback.message.edit_text(f"За {days} дн. трат не найдено.")
        return

    total = 0
    text = f"📅 <b>Отчёт за {days} дн.:</b>\n\n"
    for name, price, date in rows:
        total += price
        d_obj = datetime.strptime(str(date).split(".")[0], "%Y-%m-%d %H:%M:%S")
        text += f"• {d_obj.strftime('%d.%m')}: {name} — {price:.2f} ₽\n"
    
    text += f"\n<b>Итого: {total:.2f} ₽</b>"
    if len(text) > 4000:
        text = text[:4000] + "\n...(обрезано)"
    
    await callback.message.edit_text(text, parse_mode="HTML")

@dp.callback_query(F.data == "rep_excel")
async def report_excel(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    days = data.get("days", 7)
    await state.clear()

    await callback.message.edit_text("⏳ Генерирую Excel...")
    
    start_date = datetime.now() - timedelta(days=days)
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT name, price, date FROM expenses WHERE date >= ? ORDER BY date DESC", (start_date,)) as cur:
            rows = await cur.fetchall()

    if not rows:
        await callback.message.edit_text("Трат не найдено.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расходы"
    
    bold_font = Font(bold=True)
    ws.append(["Дата", "Товар", "Сумма"])
    for cell in ws[1]: cell.font = bold_font

    total = 0
    for name, price, date in rows:
        total += price
        d_obj = datetime.strptime(str(date).split(".")[0], "%Y-%m-%d %H:%M:%S")
        ws.append([d_obj.strftime('%d.%m.%Y %H:%M'), name, price])

    ws.append([])
    ws.append(["ИТОГО", "", total])
    ws["C" + str(ws.max_row)].font = bold_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    file = BufferedInputFile(buf.getvalue(), filename=f"report_{days}days.xlsx")
    await callback.message.answer_document(file, caption=f"📊 Итого: {total:.2f} ₽")
    await callback.message.delete()

# --- 3. Чек (Фото) + Улучшенное редактирование ---

async def send_ocr_list(message: Message, items: List[Tuple[str, float]], state: FSMContext):
    """
    Отправляет новый список и запоминает ID сообщения в состоянии.
    """
    if not items:
        await message.answer("Список пуст.", reply_markup=get_main_kb())
        await state.clear()
        return

    text = "🧾 <b>Распознано:</b>\n"
    for i, (name, price) in enumerate(items, 1):
        text += f"<b>{i}.</b> {name} — {price} ₽\n"
    
    text += f"\nИтого: {sum(x[1] for x in items):.2f} ₽"
    
    # Отправляем новое сообщение
    sent_msg = await message.answer(text, reply_markup=get_ocr_kb(), parse_mode="HTML")
    
    # Сохраняем ID сообщения со списком в состояние
    await state.update_data(list_msg_id=sent_msg.message_id)

@dp.message(F.text == "📷 Чек (фото)")
async def ask_photo(message: Message):
    await message.answer("📸 Пришли фото чека.")

@dp.message(F.photo)
async def handle_photo(message: Message, state: FSMContext):
    # Если было предыдущее состояние - очистим
    await state.clear()
    
    msg = await message.answer("🤖 Читаю чек...")
    buf = io.BytesIO()
    await bot.download(message.photo[-1], buf)
    items = await recognize_receipt_with_ai(buf.getvalue())
    await msg.delete()
    
    if not items:
        await message.answer("Не удалось прочитать.")
        return

    PENDING_GOODS[message.from_user.id] = items
    
    # Используем helper для отправки списка и сохранения состояния
    await send_ocr_list(message, items, state)
    # Переходим в "просмотр" (необязательно состояние, но полезно для контекста)
    await state.set_state(OCRState.viewing_list)

# Редактирование / Удаление строки
@dp.callback_query(F.data.in_({"ocr_edit", "ocr_delete"}))
async def ocr_modify_start(callback: CallbackQuery, state: FSMContext):
    mode = "delete" if callback.data == "ocr_delete" else "edit"
    await state.update_data(ocr_mode=mode)
    
    action = "удаления" if mode == "delete" else "редактирования"
    prompt_msg = await callback.message.answer(f"Введите <b>номер строки</b> для {action}:", parse_mode="HTML")
    
    # Запоминаем ID вопроса, чтобы потом удалить
    await state.update_data(prompt_msg_id=prompt_msg.message_id)
    await state.set_state(OCRState.waiting_for_line_number)
    await callback.answer()

@dp.message(OCRState.waiting_for_line_number)
async def ocr_line_number_input(message: Message, state: FSMContext):
    if await check_menu_break(message, state): return
    
    # 1. Получаем ID сообщений, которые нужно почистить
    data = await state.get_data()
    list_msg_id = data.get("list_msg_id")     # ID большого списка
    prompt_msg_id = data.get("prompt_msg_id") # ID вопроса "Введите номер..."
    
    # 2. Удаляем сообщение юзера (цифру) и вопрос бота
    await try_delete(bot, message.chat.id, message.message_id)
    if prompt_msg_id: await try_delete(bot, message.chat.id, prompt_msg_id)

    # Проверка на число
    if not message.text.isdigit():
        err = await message.answer("⚠️ Нужно ввести число!")
        await asyncio.sleep(2)
        await err.delete()
        return

    idx = int(message.text) - 1
    user_id = message.from_user.id
    items = PENDING_GOODS.get(user_id, [])

    # Проверка, существует ли такая строка
    if idx < 0 or idx >= len(items):
        err = await message.answer("⚠️ Нет такого номера строки.")
        await asyncio.sleep(2)
        await err.delete()
        return

    mode = data.get("ocr_mode")

    if mode == "delete":
        # --- ЛОГИКА УДАЛЕНИЯ ---
        items.pop(idx)
        PENDING_GOODS[user_id] = items
        
        # Удаляем СТАРЫЙ список
        if list_msg_id: await try_delete(bot, message.chat.id, list_msg_id)
        
        # Показываем НОВЫЙ список
        await send_ocr_list(message, items, state)
        await state.set_state(OCRState.viewing_list)
        
    else:
        # --- ЛОГИКА РЕДАКТИРОВАНИЯ ---
        await state.update_data(ocr_index=idx)
        item = items[idx]
        
        # Отправляем вопрос "Введи новое" и запоминаем его ID
        prompt = await message.answer(
            f"✏️ Редактируем строку <b>{idx+1}</b>:\n"
            f"Было: <i>{item[0]} — {item[1]}</i>\n\n"
            f"👇 Введи новые данные (формат: <b>Товар Цена</b>):", 
            parse_mode="HTML"
        )
        # Важно: перезаписываем prompt_msg_id, чтобы потом удалить именно этот вопрос
        await state.update_data(prompt_msg_id=prompt.message_id)
        
        # Переходим в режим ожидания текста
        await state.set_state(OCRState.waiting_for_new_data)

@dp.message(OCRState.waiting_for_new_data)
async def ocr_edit_save(message: Message, state: FSMContext):
    # 1. Удаляем сообщение юзера ("Кетчуп 40"), чтобы было чисто
    await try_delete(bot, message.chat.id, message.message_id)

    data = await state.get_data()
    prompt_msg_id = data.get("prompt_msg_id") # ID вопроса "Введи новые данные"
    list_msg_id = data.get("list_msg_id")     # ID старого списка

    try:
        # Парсим ввод
        parts = message.text.rsplit(maxsplit=1)
        if len(parts) < 2: raise ValueError
        name, price = parts[0].strip(), float(parts[1].replace(',', '.'))
        
        # 2. Удаляем вопрос бота "Введи новые данные..."
        if prompt_msg_id: await try_delete(bot, message.chat.id, prompt_msg_id)
        
        # 3. Обновляем данные в памяти
        idx = data['ocr_index']
        user_id = message.from_user.id
        
        if user_id in PENDING_GOODS:
            PENDING_GOODS[user_id][idx] = (name, price)
            
            # 4. --- ГЛАВНОЕ: АНАЛОГИЯ С УДАЛЕНИЕМ ---
            # Удаляем СТАРЫЙ список (чтобы не висел дубль)
            if list_msg_id: await try_delete(bot, message.chat.id, list_msg_id)

            # Отправляем НОВЫЙ список
            await send_ocr_list(message, PENDING_GOODS[user_id], state)
            
            # Возвращаем состояние просмотра
            await state.set_state(OCRState.viewing_list)
        else:
            await message.answer("⚠️ Список устарел или был очищен.")
            await state.clear()

    except ValueError:
        # Если формат кривой, ругаемся, но состояние НЕ сбрасываем
        err = await message.answer("⚠️ Ошибка формата! Пример: Кетчуп 40")
        await asyncio.sleep(3)
        await err.delete()
        
# Сохранение чека
@dp.callback_query(F.data == "ocr_save")
async def ocr_save(callback: CallbackQuery, state: FSMContext):
    items = PENDING_GOODS.pop(callback.from_user.id, [])
    
    # Очищаем состояние, чтобы бот не искал траты
    await state.clear()

    if items:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.executemany("INSERT INTO expenses (name, price) VALUES (?, ?)", items)
            await db.commit()
        # Редактируем сообщение со списком на "Сохранено"
        await callback.message.edit_text(f"✅ Успешно сохранено {len(items)} позиций.\nБот готов к работе.", reply_markup=None)
    else:
        await callback.message.delete()
        await callback.answer("Пусто")

@dp.callback_query(F.data == "ocr_cancel")
async def ocr_cancel(callback: CallbackQuery, state: FSMContext):
    PENDING_GOODS.pop(callback.from_user.id, None)
    await state.clear()
    await callback.message.delete()
    await callback.message.answer("❌ Добавление чека отменено.", reply_markup=get_main_kb())

# --- 4. Прочее (Списки, Очистка) ---

@dp.message(F.text == "📋 Список трат")
async def list_expenses(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT name, price, date FROM expenses ORDER BY date DESC LIMIT 10") as cur:
            rows = await cur.fetchall()
    text = "\n".join([f"{r[0]} — {r[1]} ₽" for r in rows]) if rows else "Пусто"
    await message.answer(f"📋 Последние 10:\n{text}")

@dp.message(F.text == "📊 Статистика")
async def stats(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT SUM(price) FROM expenses") as cur:
            res = await cur.fetchone()
    await message.answer(f"💰 Всего потрачено: {res[0] or 0} ₽")

@dp.message(F.text == "🗑 Удалить историю")
async def clear_ask(message: Message):
    await message.answer("Точно удалить всё?", reply_markup=get_confirm_delete_kb())

@dp.callback_query(F.data == "confirm_delete_history")
async def clear_confirm(callback: CallbackQuery):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM expenses")
        await db.commit()
    await callback.message.edit_text("✅ История очищена.")

async def main():
    await init_db()
    print("Бот запущен...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())